import os
from datetime import datetime, timezone, timedelta
from io import BytesIO

from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from telegram import (
    Update,
    ReplyKeyboardMarkup,
    InputFile,
)
from telegram.constants import ParseMode
from telegram.helpers import mention_html
from telegram.ext import (
    Application,
    CommandHandler,
    ContextTypes,
    ChatMemberHandler,
    MessageHandler,
    filters,
)

from psycopg_pool import ConnectionPool
from zoneinfo import ZoneInfo


# =========================
# إعدادات أساسية
# =========================
BOT_TOKEN = os.getenv("BOT_TOKEN", "").strip()
OWNER_ID = int(os.getenv("OWNER_ID", "0").strip())
DATABASE_URL = os.getenv("DATABASE_URL", "").strip()

DAILY_CHECK_HOUR_UTC = int(os.getenv("DAILY_CHECK_HOUR_UTC", "9"))
WARN_STAGES_OWNER = [7, 3, 1]
GROUP_MENTION_CHUNK_SIZE = 30
GRACE_PERIOD_DAYS = int(os.getenv("GRACE_PERIOD_DAYS", "5"))

KING_USERNAME = "@Alking03"
LOCAL_TIMEZONE = os.getenv("LOCAL_TIMEZONE", "UTC").strip()

if not BOT_TOKEN or OWNER_ID == 0 or not DATABASE_URL:
    raise SystemExit("Missing env vars. Please set BOT_TOKEN, OWNER_ID, DATABASE_URL.")


def get_tz() -> ZoneInfo:
    try:
        return ZoneInfo(LOCAL_TIMEZONE)
    except Exception:
        return ZoneInfo("UTC")


TZ_LOCAL = get_tz()


# =========================
# وقت وتنسيق
# =========================
def ensure_aware_utc(dt: datetime) -> datetime:
    if dt is None:
        return dt
    if dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def to_local(dt: datetime) -> datetime:
    dt = ensure_aware_utc(dt)
    return dt.astimezone(TZ_LOCAL)


def fmt_local(dt: datetime) -> str:
    d = to_local(dt)
    return d.strftime("%Y-%m-%d %H:%M:%S")


# =========================
# PostgreSQL Pool
# =========================
pool = ConnectionPool(conninfo=DATABASE_URL, min_size=1, max_size=5, open=True)


def init_db():
    with pool.connection() as conn:
        with conn.cursor() as cur:
            # إنشاء الجدول الأساسي إن لم يكن موجود
            cur.execute("""
            CREATE TABLE IF NOT EXISTS members (
                user_id BIGINT PRIMARY KEY,
                username TEXT,
                full_name TEXT,
                joined_at TIMESTAMPTZ NOT NULL,
                expires_at TIMESTAMPTZ NOT NULL,
                warn_stage_sent INTEGER
            );
            """)

            # ⬅️ إضافة الأعمدة الجديدة بأمان (لو القاعدة قديمة)
            cur.execute("ALTER TABLE members ADD COLUMN IF NOT EXISTS is_active BOOLEAN NOT NULL DEFAULT TRUE;")
            cur.execute("ALTER TABLE members ADD COLUMN IF NOT EXISTS left_at TIMESTAMPTZ NULL;")

            # ⬅️ إنشاء الـ indexes بعد التأكد من وجود الأعمدة
            cur.execute("CREATE INDEX IF NOT EXISTS idx_members_expires_at ON members (expires_at);")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_members_active ON members (is_active);")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_members_active_expires ON members (is_active, expires_at);")

            # إعدادات البوت
            cur.execute("""
            CREATE TABLE IF NOT EXISTS settings (
                key TEXT PRIMARY KEY,
                value TEXT
            );
            """)

            cur.execute("""
            INSERT INTO settings(key, value) VALUES
                ('default_duration_months', '2'),
                ('notify_group_chat_id', ''),
                ('group_notify_enabled', '0'),
                ('group_notify_stages', '7')
            ON CONFLICT (key) DO NOTHING;
            """)

            # لو قاعدة البيانات قديمة وكانت مدة الاشتراك الافتراضية 3 شهور،
            # غيّرها تلقائيًا إلى شهرين بدون ما يمس باقي الإعدادات.
            cur.execute("""
            UPDATE settings
            SET value='2'
            WHERE key='default_duration_months' AND value='3';
            """)

        conn.commit()



def get_setting(key: str, default: str = "") -> str:
    with pool.connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT value FROM settings WHERE key=%s", (key,))
            row = cur.fetchone()
            return row[0] if row and row[0] is not None else default


def set_setting(key: str, value: str):
    with pool.connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
            INSERT INTO settings(key, value) VALUES(%s, %s)
            ON CONFLICT (key) DO UPDATE SET value=EXCLUDED.value
            """, (key, value))
        conn.commit()


def get_default_months() -> int:
    try:
        return int(get_setting("default_duration_months", "2"))
    except Exception:
        return 2


def compute_expiry(joined_at: datetime) -> datetime:
    return joined_at + relativedelta(months=+get_default_months())


def upsert_member_join(user_id: int, username: str, full_name: str, joined_at: datetime):
    """
    عند الانضمام (حتى لو كان قديم ورجع):
    - نسجل joined_at الجديد
    - نحسب expires_at جديد
    - نخليه Active
    - نمسح left_at
    - warn_stage_sent = NULL
    """
    joined_at = ensure_aware_utc(joined_at)
    expires = compute_expiry(joined_at)
    with pool.connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
            INSERT INTO members(user_id, username, full_name, joined_at, expires_at, warn_stage_sent, is_active, left_at)
            VALUES (%s, %s, %s, %s, %s, NULL, TRUE, NULL)
            ON CONFLICT (user_id) DO UPDATE SET
                username=EXCLUDED.username,
                full_name=EXCLUDED.full_name,
                joined_at=EXCLUDED.joined_at,
                expires_at=EXCLUDED.expires_at,
                warn_stage_sent=NULL,
                is_active=TRUE,
                left_at=NULL
            """, (user_id, username or "", full_name or "", joined_at, expires))
        conn.commit()


def mark_member_left(user_id: int, left_at: datetime):
    """
    عند الخروج أو الطرد:
    - نخليه Removed (is_active=false)
    - نسجل left_at
    - (ما بنحذف الاشتراك من القاعدة)
    """
    left_at = ensure_aware_utc(left_at)
    with pool.connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
            UPDATE members
            SET is_active=FALSE, left_at=%s
            WHERE user_id=%s
            """, (left_at, user_id))
        conn.commit()

def mark_member_active(user_id: int):
    """رجّع العضو Active بدون ما نغيّر joined/expires (فقط حالة)."""
    with pool.connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
            UPDATE members
            SET is_active=TRUE, left_at=NULL
            WHERE user_id=%s
            """, (user_id,))
        conn.commit()


def get_counts_active() -> int:
    with pool.connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM members WHERE is_active=TRUE")
            return int(cur.fetchone()[0])


def fetch_active_all():
    with pool.connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
            SELECT user_id, username, full_name, joined_at, expires_at, warn_stage_sent, is_active, left_at
            FROM members
            WHERE is_active=TRUE
            ORDER BY joined_at DESC
            """)
            return cur.fetchall()


def fetch_removed_all():
    with pool.connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
            SELECT user_id, username, full_name, joined_at, expires_at, warn_stage_sent, is_active, left_at
            FROM members
            WHERE is_active=FALSE
            ORDER BY left_at DESC NULLS LAST
            """)
            return cur.fetchall()


def fetch_active_expired(now: datetime):
    now = ensure_aware_utc(now)
    with pool.connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
            SELECT user_id, username, full_name, joined_at, expires_at, warn_stage_sent, is_active, left_at
            FROM members
            WHERE is_active=TRUE AND expires_at < %s
            ORDER BY expires_at ASC
            """, (now,))
            return cur.fetchall()




def fetch_active_past_grace(now: datetime, grace_days: int = GRACE_PERIOD_DAYS):
    """الأعضاء الـ Active الذين انتهت مهلة السماح الخاصة بهم ويجب إزالتهم من الكروب."""
    now = ensure_aware_utc(now)
    cutoff = now - timedelta(days=grace_days)
    with pool.connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
            SELECT user_id, username, full_name, joined_at, expires_at, warn_stage_sent, is_active, left_at
            FROM members
            WHERE is_active=TRUE AND expires_at <= %s
            ORDER BY expires_at ASC
            """, (cutoff,))
            return cur.fetchall()

def fetch_active_expiring_within(now: datetime, days: int):
    now = ensure_aware_utc(now)
    end = now + timedelta(days=days)
    with pool.connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
            SELECT user_id, username, full_name, joined_at, expires_at, warn_stage_sent, is_active, left_at
            FROM members
            WHERE is_active=TRUE AND expires_at >= %s AND expires_at <= %s
            ORDER BY expires_at ASC
            """, (now, end))
            return cur.fetchall()


def find_member(query: str):
    with pool.connection() as conn:
        with conn.cursor() as cur:
            if query.isdigit():
                cur.execute("""
                SELECT user_id, username, full_name, joined_at, expires_at, warn_stage_sent, is_active, left_at
                FROM members WHERE user_id=%s
                """, (int(query),))
                return cur.fetchone()

            q = query.lstrip("@")
            cur.execute("""
            SELECT user_id, username, full_name, joined_at, expires_at, warn_stage_sent, is_active, left_at
            FROM members WHERE lower(username)=lower(%s)
            """, (q,))
            return cur.fetchone()


def extend_member_days(user_id: int, days: int) -> bool:
    with pool.connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT expires_at FROM members WHERE user_id=%s", (user_id,))
            row = cur.fetchone()
            if not row:
                return False
            old_exp = ensure_aware_utc(row[0])
            new_exp = old_exp + timedelta(days=days)
            cur.execute("""
            UPDATE members
            SET expires_at=%s, warn_stage_sent=NULL
            WHERE user_id=%s
            """, (new_exp, user_id))
        conn.commit()
    return True


def set_warn_stage(user_id: int, stage: int | None):
    with pool.connection() as conn:
        with conn.cursor() as cur:
            cur.execute("UPDATE members SET warn_stage_sent=%s WHERE user_id=%s", (stage, user_id))
        conn.commit()


# =========================
# صلاحيات المالك
# =========================
def is_owner(update: Update) -> bool:
    u = update.effective_user
    return bool(u and u.id == OWNER_ID)


async def owner_only(update: Update) -> bool:
    if is_owner(update):
        return True
    if update.message:
        await update.message.reply_text("هذا البوت مخصص للمالك فقط.")
    return False


# =========================
# Reply Keyboard UI
# =========================
STATE_KEY = "ui_state"
WAITING_SEARCH = "waiting_search"
WAITING_EXTEND_ID = "waiting_extend_id"
WAITING_SYNC = "waiting_sync"


STATE_MAIN = "MAIN"
STATE_GROUP = "GROUP"
STATE_DURATION = "DURATION"


def kb_main():
    return ReplyKeyboardMarkup(
        [
            ["📊 عدد المشتركين", "🧾 تصدير Excel"],
            ["🔎 بحث عن مشترك", "➕ تمديد +90 يوم"],
            ["🔄 تحديث حالة عضو", "🔔 فحص التنبيه الآن"],
            ["🚫 فحص الطرد بعد المهلة"],
            ["📣 إعدادات تنبيه الكروب", "⏳ مدة الاشتراك الافتراضية"],
        ],
        resize_keyboard=True
    )



def kb_group():
    return ReplyKeyboardMarkup(
        [
            ["✅ تفعيل/إيقاف تنبيه الكروب"],
            ["مراحل التنبيه: 7 فقط", "مراحل التنبيه: 7+3+1"],
            ["📌 شرح setgroup"],
            ["⬅️ رجوع للقائمة الرئيسية"],
        ],
        resize_keyboard=True
    )


def kb_duration():
    return ReplyKeyboardMarkup(
        [
            ["1 شهر", "2 شهور", "12 شهر"],
            ["⬅️ رجوع للقائمة الرئيسية"],
        ],
        resize_keyboard=True
    )


def group_status_text() -> str:
    enabled = get_setting("group_notify_enabled", "0") == "1"
    gid = get_setting("notify_group_chat_id", "").strip()
    stages = get_setting("group_notify_stages", "7").strip() or "7"
    return (
        f"تنبيه الكروب: {'مفعّل' if enabled else 'موقّف'}\n"
        f"الكروب المحفوظ: {gid if gid else '(غير محدد)'}\n"
        f"مراحل التنبيه بالكروب: {stages}\n"
        f"توقيت العرض: {LOCAL_TIMEZONE}"
    )


# =========================
# /start
# =========================
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await owner_only(update):
        return

    context.user_data[STATE_KEY] = STATE_MAIN
    context.user_data[WAITING_SEARCH] = False
    context.user_data[WAITING_EXTEND_ID] = False
    context.user_data[WAITING_SYNC] = False

    await update.message.reply_text("لوحة التحكم ✅", reply_markup=kb_main())


# =========================
# /setgroup داخل الكروب
# =========================
async def setgroup(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await owner_only(update):
        return

    chat = update.effective_chat
    if chat.type not in ("group", "supergroup"):
        await update.message.reply_text("استخدم الأمر داخل الكروب فقط.")
        return

    set_setting("notify_group_chat_id", str(chat.id))
    await update.message.reply_text("✅ تم حفظ هذا الكروب لإرسال تنبيهات الاشتراك.")

    if get_setting("group_notify_enabled", "0") != "1":
        set_setting("group_notify_enabled", "1")
        await update.message.reply_text("✅ تم تفعيل تنبيه الكروب تلقائيًا.")


# =========================
# تتبع دخول/خروج الأعضاء
# =========================
async def on_chat_member(update: Update, context: ContextTypes.DEFAULT_TYPE):
    cmu = update.chat_member
    if not cmu:
        return

    # العضو المتأثر بالحدث، وليس الشخص الذي قام بتنفيذ الحدث
    user = cmu.new_chat_member.user
    new_status = cmu.new_chat_member.status
    old_status = cmu.old_chat_member.status

    # حالات الانضمام
    joined = (old_status in ("left", "kicked")) and (new_status in ("member", "restricted", "administrator"))
    if joined:
        joined_at = datetime.now(timezone.utc)
        full_name = " ".join([p for p in [user.first_name, user.last_name] if p]) or ""
        username = user.username or ""
        upsert_member_join(user.id, username, full_name, joined_at)
        return

    # حالات الخروج/الطرد
    left_or_kicked = (old_status in ("member", "restricted", "administrator")) and (new_status in ("left", "kicked"))
    if left_or_kicked:
        left_at = datetime.now(timezone.utc)
        mark_member_left(user.id, left_at)
        return


# =========================
# Excel Export (Active/Removed)
# =========================
def autosize(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)


def row_for_excel(r):
    # r: user_id, username, full_name, joined_at, expires_at, warn_stage, is_active, left_at
    user_id, username, full_name, joined_at, expires_at, warn_stage, is_active, left_at = r
    return [
        user_id,
        username,
        full_name,
        fmt_local(joined_at),
        fmt_local(expires_at),
        warn_stage,
        bool(is_active),
        (fmt_local(left_at) if left_at else ""),
    ]


def add_sheet(wb: Workbook, title: str, rows):
    ws = wb.create_sheet(title=title)
    headers = [
        "user_id", "username", "full_name",
        "joined_at_local", "expires_at_local",
        "warn_stage_sent", "is_active", "left_at_local"
    ]
    ws.append(headers)
    for r in rows:
        ws.append(row_for_excel(r))
    autosize(ws)


def build_xlsx_bytes(active_rows, expiring_rows, expired_rows, removed_rows):
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Active"
    headers = [
        "user_id", "username", "full_name",
        "joined_at_local", "expires_at_local",
        "warn_stage_sent", "is_active", "left_at_local"
    ]
    ws0.append(headers)
    for r in active_rows:
        ws0.append(row_for_excel(r))
    autosize(ws0)

    add_sheet(wb, "Expiring_7_Days", expiring_rows)
    add_sheet(wb, "Expired", expired_rows)
    add_sheet(wb, "Removed", removed_rows)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


# =========================
# تنبيهات (Active فقط)
# =========================
def stage_for_remaining_days(remaining_days: int) -> int | None:
    if remaining_days <= 1:
        return 1
    if remaining_days <= 3:
        return 3
    if remaining_days <= 7:
        return 7
    return None


def parse_group_stages() -> set[int]:
    raw = (get_setting("group_notify_stages", "7") or "7").strip()
    parts = [p.strip() for p in raw.split(",") if p.strip()]
    out = set()
    for p in parts:
        if p.isdigit():
            out.add(int(p))
    return (out & {7, 3, 1}) if out else {7}


async def notify_group_mentions(context: ContextTypes.DEFAULT_TYPE, stage: int, people):
    enabled = get_setting("group_notify_enabled", "0") == "1"
    gid = get_setting("notify_group_chat_id", "").strip()
    if (not enabled) or (not gid):
        return

    allowed = parse_group_stages()
    if stage not in allowed:
        return

    group_id = int(gid)

    mentions = []
    for user_id, full_name in people:
        display = full_name if full_name else str(user_id)
        mentions.append(mention_html(user_id, display))

    title = f"🔔 تنبيه: باقي {stage} أيام على انتهاء الاشتراك"
    subtitle = (
        "إذا كان لديك أي استفسار أو كنت تريد تجديد اشتراكك\n"
        f"يرجى مراسلة الكينغ لتجديد الاشتراك: {KING_USERNAME}\n"
        "وشكرًا."
    )

    for i in range(0, len(mentions), GROUP_MENTION_CHUNK_SIZE):
        chunk = mentions[i:i + GROUP_MENTION_CHUNK_SIZE]
        text = title + "\n\n" + subtitle + "\n\n" + "\n".join(f"• {m}" for m in chunk)
        await context.bot.send_message(
            chat_id=group_id,
            text=text,
            parse_mode=ParseMode.HTML,
            disable_web_page_preview=True,
        )


async def run_warning_check(context: ContextTypes.DEFAULT_TYPE, manual: bool = False):
    now = datetime.now(timezone.utc)
    candidates = fetch_active_expiring_within(now, 7)

    to_owner = {7: [], 3: [], 1: []}
    to_group = {7: [], 3: [], 1: []}

    for user_id, username, full_name, _joined_at, expires_at, warn_stage_sent, _is_active, _left_at in candidates:
        expires_at = ensure_aware_utc(expires_at)
        remaining_seconds = (expires_at - ensure_aware_utc(now)).total_seconds()
        remaining_days = int((remaining_seconds + 86399) // 86400)

        stage = stage_for_remaining_days(remaining_days)
        if stage is None:
            continue

        if warn_stage_sent is not None and int(warn_stage_sent) <= stage:
            continue

        u = f"@{username}" if username else "(بدون يوزرنيم)"
        exp_local = fmt_local(expires_at)

        to_owner[stage].append((user_id, full_name, u, exp_local, remaining_days))
        to_group[stage].append((user_id, full_name))

    sent_any = False

    for stage in WARN_STAGES_OWNER:
        items = to_owner[stage]
        if not items:
            continue

        sent_any = True
        lines = [f"🔔 تنبيه قبل {stage} أيام (توقيتك: {LOCAL_TIMEZONE})", ""]
        for user_id, full_name, u, exp_local, remaining_days in items[:60]:
            lines.append(f"- {full_name} | {u} | ID: {user_id} | ينتهي: {exp_local} | باقي: {remaining_days} يوم")
        if len(items) > 60:
            lines.append("")
            lines.append(f"… ويوجد {len(items) - 60} آخرين (صدّر Excel لرؤية الجميع).")

        await context.bot.send_message(chat_id=OWNER_ID, text="\n".join(lines))
        await notify_group_mentions(context, stage, to_group[stage])

        for user_id, *_ in items:
            set_warn_stage(user_id, stage)

    if manual and not sent_any:
        await context.bot.send_message(chat_id=OWNER_ID, text="✅ لا يوجد تنبيهات الآن.")




async def kick_member_after_grace(context: ContextTypes.DEFAULT_TYPE, group_id: int, row) -> tuple[bool, str]:
    user_id, username, full_name, _joined_at, expires_at, _warn_stage, _is_active, _left_at = row
    display_name = full_name or (f"@{username}" if username else str(user_id))

    try:
        cm = await context.bot.get_chat_member(chat_id=group_id, user_id=user_id)
        if cm.status in ("left", "kicked"):
            mark_member_left(user_id, datetime.now(timezone.utc))
            return True, f"- {display_name} | ID: {user_id} | كان خارج الكروب بالفعل وتم ضبطه Removed"

        if cm.status in ("administrator", "creator"):
            return False, f"- {display_name} | ID: {user_id} | لم يتم طرده لأنه Admin/Owner"

        # ban + unban = طرد فقط بدون حظر دائم
        await context.bot.ban_chat_member(chat_id=group_id, user_id=user_id)
        await context.bot.unban_chat_member(chat_id=group_id, user_id=user_id, only_if_banned=True)
        mark_member_left(user_id, datetime.now(timezone.utc))

        grace_ended = ensure_aware_utc(expires_at) + timedelta(days=GRACE_PERIOD_DAYS)
        return True, f"- {display_name} | ID: {user_id} | انتهت المهلة: {fmt_local(grace_ended)}"

    except Exception as e:
        return False, f"- {display_name} | ID: {user_id} | فشل الطرد: {type(e).__name__}"


async def run_grace_kick_check(context: ContextTypes.DEFAULT_TYPE, manual: bool = False):
    gid = get_setting("notify_group_chat_id", "").strip()
    if not gid:
        if manual:
            await context.bot.send_message(
                chat_id=OWNER_ID,
                text="ما في كروب محفوظ للطرد. نفّذ /setgroup داخل الكروب أولًا."
            )
        return

    group_id = int(gid)
    now = datetime.now(timezone.utc)
    expired_after_grace = fetch_active_past_grace(now, GRACE_PERIOD_DAYS)

    if not expired_after_grace:
        if manual:
            await context.bot.send_message(chat_id=OWNER_ID, text="✅ لا يوجد أعضاء تجاوزوا مهلة السماح الآن.")
        return

    success_lines = []
    failed_lines = []

    for row in expired_after_grace:
        ok, line = await kick_member_after_grace(context, group_id, row)
        if ok:
            success_lines.append(line)
        else:
            failed_lines.append(line)

    lines = [f"🚫 فحص إزالة المنتهية اشتراكاتهم بعد مهلة {GRACE_PERIOD_DAYS} أيام", ""]
    if success_lines:
        lines.append("تمت الإزالة / التحديث:")
        lines.extend(success_lines[:80])
        if len(success_lines) > 80:
            lines.append(f"… ويوجد {len(success_lines) - 80} آخرين.")

    if failed_lines:
        if success_lines:
            lines.append("")
        lines.append("لم تتم إزالة هؤلاء:")
        lines.extend(failed_lines[:80])
        if len(failed_lines) > 80:
            lines.append(f"… ويوجد {len(failed_lines) - 80} آخرين.")

    await context.bot.send_message(chat_id=OWNER_ID, text="\n".join(lines))

async def job_daily_warning(context: ContextTypes.DEFAULT_TYPE):
    await run_warning_check(context, manual=False)
    await run_grace_kick_check(context, manual=False)


# =========================
# نصوص الأزرار
# =========================
TXT_COUNT = "📊 عدد المشتركين"
TXT_EXPORT = "🧾 تصدير Excel"
TXT_SEARCH = "🔎 بحث عن مشترك"
TXT_EXTEND = "➕ تمديد +90 يوم"
TXT_WARN = "🔔 فحص التنبيه الآن"
TXT_KICK_GRACE = "🚫 فحص الطرد بعد المهلة"
TXT_GROUP = "📣 إعدادات تنبيه الكروب"
TXT_DURATION = "⏳ مدة الاشتراك الافتراضية"
TXT_SYNC = "🔄 تحديث حالة عضو"

TXT_TOGGLE_GROUP = "✅ تفعيل/إيقاف تنبيه الكروب"
TXT_STAGE_7 = "مراحل التنبيه: 7 فقط"
TXT_STAGE_ALL = "مراحل التنبيه: 7+3+1"
TXT_HOW_SETGROUP = "📌 شرح setgroup"
TXT_BACK = "⬅️ رجوع للقائمة الرئيسية"


# =========================
# استقبال ضغط الأزرار كنص
# =========================
async def on_private_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await owner_only(update):
        return

    text = (update.message.text or "").strip()

    # ينتظر بحث
    if context.user_data.get(WAITING_SEARCH):
        context.user_data[WAITING_SEARCH] = False
        row = find_member(text)
        if not row:
            await update.message.reply_text("لم يتم العثور على هذا المشترك.", reply_markup=kb_main())
            return

        user_id, username, full_name, joined_at, expires_at, warn_stage, is_active, left_at = row
        u = f"@{username}" if username else "(لا يوجد)"
        status = "Active ✅" if is_active else "Removed ❌"
        msg = (
            "✅ بيانات المشترك:\n"
            f"- ID: {user_id}\n"
            f"- Username: {u}\n"
            f"- Name: {full_name}\n"
            f"- Status: {status}\n"
            f"- Joined: {fmt_local(joined_at)} ({LOCAL_TIMEZONE})\n"
            f"- Expires: {fmt_local(expires_at)} ({LOCAL_TIMEZONE})\n"
            f"- Grace ends: {fmt_local(ensure_aware_utc(expires_at) + timedelta(days=GRACE_PERIOD_DAYS))} ({LOCAL_TIMEZONE})\n"
            f"- Left at: {(fmt_local(left_at) if left_at else '-')}\n"
            f"- Warn stage sent: {warn_stage if warn_stage is not None else 'None'}"
        )
        await update.message.reply_text(msg, reply_markup=kb_main())
        return

    # ينتظر تمديد
    if context.user_data.get(WAITING_EXTEND_ID):
        context.user_data[WAITING_EXTEND_ID] = False
        if not text.isdigit():
            await update.message.reply_text("أرسل رقم ID فقط.", reply_markup=kb_main())
            return
        ok = extend_member_days(int(text), 90)
        if not ok:
            await update.message.reply_text("هذا الـID غير موجود في قاعدة البيانات.", reply_markup=kb_main())
            return
        await update.message.reply_text("✅ تم التمديد +90 يوم (3 شهور).", reply_markup=kb_main())
        return

    # ينتظر مزامنة حالة عضو
    if context.user_data.get(WAITING_SYNC):
        context.user_data[WAITING_SYNC] = False

        if not text.isdigit():
            await update.message.reply_text("أرسل رقم ID فقط.", reply_markup=kb_main())
            return

        gid = get_setting("notify_group_chat_id", "").strip()
        if not gid:
            await update.message.reply_text("ما في كروب محفوظ. نفّذ /setgroup داخل الكروب أولًا.", reply_markup=kb_main())
            return

        group_id = int(gid)
        user_id = int(text)

        try:
            cm = await context.bot.get_chat_member(chat_id=group_id, user_id=user_id)
            status = cm.status

            if status in ("left", "kicked"):
                mark_member_left(user_id, datetime.now(timezone.utc))
                await update.message.reply_text(f"✅ تم ضبط العضو كـ Removed (status={status}).", reply_markup=kb_main())
            else:
                mark_member_active(user_id)
                await update.message.reply_text(f"✅ تم ضبط العضو كـ Active (status={status}).", reply_markup=kb_main())

        except Exception:
            await update.message.reply_text(
                "تعذّر فحص العضو.\n"
                "تأكد من:\n"
                "- البوت Admin بالكروب\n"
                "- نفذت /setgroup\n"
                "- الـID صحيح",
                reply_markup=kb_main()
            )

        return

    state = context.user_data.get(STATE_KEY, STATE_MAIN)

    # رجوع
    if text == TXT_BACK:
        context.user_data[STATE_KEY] = STATE_MAIN
        await update.message.reply_text("✅ رجعنا للقائمة الرئيسية", reply_markup=kb_main())
        return

    # MAIN
    if state == STATE_MAIN:
        if text == TXT_COUNT:
            await update.message.reply_text(f"📊 عدد المشتركين (Active): {get_counts_active()}", reply_markup=kb_main())
            return

        if text == TXT_EXPORT:
            now = datetime.now(timezone.utc)
            bio = build_xlsx_bytes(
                active_rows=fetch_active_all(),
                expiring_rows=fetch_active_expiring_within(now, 7),
                expired_rows=fetch_active_expired(now),
                removed_rows=fetch_removed_all(),
            )
            bio.name = "subscribers.xlsx"
            await update.message.reply_document(document=InputFile(bio), caption="ملف Excel ✅", reply_markup=kb_main())
            return

        if text == TXT_SEARCH:
            context.user_data[WAITING_SEARCH] = True
            await update.message.reply_text("أرسل رقم الـID أو اليوزرنيم (مثال: 123456 أو @username).", reply_markup=kb_main())
            return

        if text == TXT_EXTEND:
            context.user_data[WAITING_EXTEND_ID] = True
            await update.message.reply_text("أرسل رقم ID للمشترك لتمديده +90 يوم (3 شهور).", reply_markup=kb_main())
            return

        if text == TXT_SYNC:
            context.user_data[WAITING_SYNC] = True
            await update.message.reply_text("أرسل رقم ID للعضو لمزامنة حالته مع الكروب (Active / Removed).", reply_markup=kb_main())
            return

        if text == TXT_WARN:
            await run_warning_check(context, manual=True)
            await update.message.reply_text("✅ تم فحص التنبيهات الآن.", reply_markup=kb_main())
            return

        if text == TXT_KICK_GRACE:
            await run_grace_kick_check(context, manual=True)
            await update.message.reply_text("✅ تم فحص الطرد بعد المهلة الآن.", reply_markup=kb_main())
            return

        if text == TXT_GROUP:
            context.user_data[STATE_KEY] = STATE_GROUP
            await update.message.reply_text(group_status_text(), reply_markup=kb_group())
            return

        if text == TXT_DURATION:
            context.user_data[STATE_KEY] = STATE_DURATION
            await update.message.reply_text("اختر مدة الاشتراك الافتراضية للأعضاء الجدد:", reply_markup=kb_duration())
            return

        await update.message.reply_text("اختر من الأزرار الموجودة بالأسفل 👇", reply_markup=kb_main())
        return

    # GROUP
    if state == STATE_GROUP:
        if text == TXT_TOGGLE_GROUP:
            cur = get_setting("group_notify_enabled", "0")
            set_setting("group_notify_enabled", "0" if cur == "1" else "1")
            await update.message.reply_text(group_status_text(), reply_markup=kb_group())
            return

        if text == TXT_STAGE_7:
            set_setting("group_notify_stages", "7")
            await update.message.reply_text(group_status_text(), reply_markup=kb_group())
            return

        if text == TXT_STAGE_ALL:
            set_setting("group_notify_stages", "7,3,1")
            await update.message.reply_text(group_status_text(), reply_markup=kb_group())
            return

        if text == TXT_HOW_SETGROUP:
            await update.message.reply_text(
                "لتحديد الكروب الذي تُرسل إليه التنبيهات:\n"
                "1) أضف البوت كـ Admin بالكروب\n"
                "2) داخل الكروب اكتب: /setgroup\n"
                "بعدها فعّل تنبيه الكروب من هنا.",
                reply_markup=kb_group()
            )
            return

        await update.message.reply_text("اختر من خيارات إعدادات الكروب 👇", reply_markup=kb_group())
        return

    # DURATION
    if state == STATE_DURATION:
        if text in ("1 شهر", "2 شهور", "12 شهر"):
            months = {"1 شهر": "1", "2 شهور": "2", "12 شهر": "12"}[text]
            set_setting("default_duration_months", months)
            context.user_data[STATE_KEY] = STATE_MAIN
            await update.message.reply_text(f"✅ تم ضبط المدة الافتراضية إلى: {months} شهر", reply_markup=kb_main())
            return

        await update.message.reply_text("اختر مدة من الأزرار 👇", reply_markup=kb_duration())
        return



# =========================
# تشغيل
# =========================
def main():
    init_db()

    app = Application.builder().token(BOT_TOKEN).build()

    app.add_handler(CommandHandler("start", start, filters=filters.ChatType.PRIVATE))
    app.add_handler(CommandHandler("setgroup", setgroup))
    app.add_handler(MessageHandler(filters.TEXT & filters.ChatType.PRIVATE, on_private_text))

    # تتبع دخول/خروج
    app.add_handler(ChatMemberHandler(on_chat_member, ChatMemberHandler.CHAT_MEMBER))
    app.add_handler(ChatMemberHandler(on_chat_member, ChatMemberHandler.MY_CHAT_MEMBER))


    # فحص يومي
    app.job_queue.run_daily(
        job_daily_warning,
        time=datetime.now(timezone.utc).replace(hour=DAILY_CHECK_HOUR_UTC, minute=0, second=0, microsecond=0).time(),
        name="daily_warning_check",
    )

    app.run_polling(allowed_updates=["message", "chat_member", "my_chat_member"])


if __name__ == "__main__":
    main()




